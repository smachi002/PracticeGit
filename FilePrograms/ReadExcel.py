import openpyxl

def readText(Path):
    workbook = openpyxl.load_workbook(Path)
    sheet = workbook['Attendance']
    row = sheet.max_row
    column = sheet.max_column
    for i in range(1,row+1):
        for j in range(1,column+1):
            print(sheet.cell(i,j).value, end= " ")
        print('')


if __name__ == '__main__':
    Path = '.\\Test.xlsx'
    readText(Path)
